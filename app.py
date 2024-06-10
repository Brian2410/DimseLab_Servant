from flask import Flask, request, render_template_string, redirect, url_for, send_file
import openpyxl
import qrcode
import io
import socket

app = Flask(__name__)

# Path to your Excel file
EXCEL_FILE = 'inventory.xlsx'

# Load inventory
def load_inventory():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    return wb

# Save inventory
def save_inventory(wb):
    wb.save(EXCEL_FILE)

# Generate QR code
def generate_qr_code(data):
    url = f'https://brian2410.github.io/DimseLab_Servant/inventory.html?item={data}'
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill='black', back_color='white')
    return img

# Route to generate and display QR code
@app.route('/generate_qr/<item_id>')
def generate_qr(item_id):
    img = generate_qr_code(item_id)
    img_io = io.BytesIO()
    img.save(img_io, 'PNG')
    img_io.seek(0)
    return send_file(img_io, mimetype='image/png')

# Route to display inventory and generate QR codes
@app.route('/')
def index():
    wb = load_inventory()
    sheet = wb.active
    inventory = [(row[0].value, row[1].value, row[2].value) for row in sheet.iter_rows(min_row=2)]
    return render_template_string('''
        <h1>Inventory</h1>
        <ul>
            {% for item_id, item_name, quantity in inventory %}
            <li>{{ item_name }} ({{ quantity }}) - <a href="{{ url_for('generate_qr', item_id=item_id) }}">Generate QR</a></li>
            {% endfor %}
        </ul>
        <a href="{{ url_for('inventory') }}">View Inventory HTML</a>
    ''', inventory=inventory)

# Route to handle renting
@app.route('/rent/<item_id>', methods=['GET', 'POST'])
def rent(item_id):
    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        wb = load_inventory()
        sheet = wb.active
        for row in sheet.iter_rows(min_row=2):
            if row[0].value == item_id:
                row[2].value -= 1  # Deduct one from quantity
                break
        save_inventory(wb)
        return redirect(url_for('index'))
    return '''
        <form method="post">
            Name: <input type="text" name="name"><br>
            Phone: <input type="text" name="phone"><br>
            <input type="submit" value="Rent">
        </form>
    '''

# Route to redirect to inventory HTML
@app.route('/inventory')
def inventory():
    return redirect('https://brian2410.github.io/DimseLab_Servant/inventory.html')

if __name__ == '__main__':
    app.run(debug=True, port=5000)
